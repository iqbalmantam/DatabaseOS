from datetime import date
import io
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from streamlit_gsheets import GSheetsConnection

# Set Halaman Streamlit
st.set_page_config(
    page_title="Employee Database Manager", page_icon="👥", layout="wide"
)

# --- SEMBUNYIKAN MENU KANAN ATAS ---
st.markdown(
    """
    <style>
    div[data-testid="stToolbarActions"] { display: none !important; }
    div[data-testid="stDecoration"] { display: none !important; }
    #MainMenu { visibility: hidden !important; }
    [data-testid="stCollapsedControl"] { display: flex !important; visibility: visible !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

# --- PIN ADMINISTRATOR ---
ADMIN_PIN = "2273"

# --- KONEKSI GOOGLE SHEETS ---
conn = st.connection("gsheets", type=GSheetsConnection)


# Fungsi Membaca Data Master
def load_data():
  try:
    df = conn.read(worksheet="Master_Karyawan", ttl=0)
    if df is not None and not df.empty:
      if "ID" in df.columns:
        df["ID"] = df["ID"].astype(str)
      if "Jabatan" in df.columns and "Posisi" not in df.columns:
        df.rename(columns={"Jabatan": "Posisi"}, inplace=True)
      if "Site" not in df.columns:
        df["Site"] = ""
      if "Status" not in df.columns:
        df["Status"] = "Aktif"
      if "Tanggal Resign" not in df.columns:
        df["Tanggal Resign"] = "-"
      if "Terakhir Diperbarui" not in df.columns:
        df["Terakhir Diperbarui"] = str(date.today())
    return df
  except Exception:
    try:
      df = conn.read(ttl=0)
      if df is not None and not df.empty:
        if "ID" in df.columns:
          df["ID"] = df["ID"].astype(str)
        if "Jabatan" in df.columns and "Posisi" not in df.columns:
          df.rename(columns={"Jabatan": "Posisi"}, inplace=True)
        if "Site" not in df.columns:
          df["Site"] = ""
        if "Status" not in df.columns:
          df["Status"] = "Aktif"
        if "Tanggal Resign" not in df.columns:
          df["Tanggal Resign"] = "-"
        if "Terakhir Diperbarui" not in df.columns:
          df["Terakhir Diperbarui"] = str(date.today())
      return df
    except Exception as e:
      st.error(f"Gagal terhubung ke Google Sheets: {e}")
      return pd.DataFrame(
          columns=[
              "ID",
              "Nama Lengkap",
              "Posisi",
              "Cost Center",
              "Tanggal Bergabung",
              "Akhir Kontrak",
              "Tanggal Resign",
              "Site",
              "Status",
              "Terakhir Diperbarui",
          ]
      )


# Fungsi Membaca Data Snapshot Bulanan
def load_snapshot_data():
  try:
    df_snap = conn.read(worksheet="Snapshot_Bulanan", ttl=0)
    return df_snap
  except Exception:
    return pd.DataFrame()


# Fungsi Menyimpan Data Master
def save_data(df):
  conn.update(worksheet="Master_Karyawan", data=df)
  st.session_state.employees = df


# Generator PDF
def generate_pdf(df):
  pdf = FPDF(orientation="L", unit="mm", format="A4")
  pdf.add_page()
  pdf.set_font("Helvetica", "B", 14)

  pdf.cell(0, 10, "LAPORAN DATABASE KARYAWAN", ln=True, align="C")
  pdf.set_font("Helvetica", "", 10)
  pdf.cell(
      0,
      6,
      f"Dicetak Tanggal: {date.today().strftime('%d-%m-%Y')} | Total"
      f" Record: {len(df)}",
      ln=True,
      align="C",
  )
  pdf.ln(4)

  pdf.set_font("Helvetica", "B", 8)
  pdf.set_fill_color(230, 230, 230)

  col_widths = [22, 45, 35, 25, 25, 25, 25, 25, 20, 30]
  headers = [
      "ID",
      "Nama Lengkap",
      "Posisi",
      "Cost Center",
      "Tgl Join",
      "End Kontrak",
      "Tgl Resign",
      "Site",
      "Status",
      "Updated",
  ]

  for i, h in enumerate(headers):
    pdf.cell(col_widths[i], 7, h, border=1, align="C", fill=True)
  pdf.ln()

  pdf.set_font("Helvetica", "", 7)
  for _, row in df.iterrows():
    pdf.cell(col_widths[0], 6, str(row.get("ID", "")), border=1, align="C")
    pdf.cell(col_widths[1], 6, str(row.get("Nama Lengkap", ""))[:25], border=1)
    pdf.cell(col_widths[2], 6, str(row.get("Posisi", ""))[:20], border=1)
    pdf.cell(
        col_widths[3], 6, str(row.get("Cost Center", "")), border=1, align="C"
    )
    pdf.cell(
        col_widths[4],
        6,
        str(row.get("Tanggal Bergabung", "")),
        border=1,
        align="C",
    )
    pdf.cell(
        col_widths[5], 6, str(row.get("Akhir Kontrak", "")), border=1, align="C"
    )
    pdf.cell(
        col_widths[6], 6, str(row.get("Tanggal Resign", "-")), border=1, align="C"
    )
    pdf.cell(col_widths[7], 6, str(row.get("Site", "")), border=1, align="C")
    pdf.cell(
        col_widths[8],
        6,
        str(row.get("Status", "Aktif")),
        border=1,
        align="C",
    )
    pdf.cell(
        col_widths[9],
        6,
        str(row.get("Terakhir Diperbarui", "")),
        border=1,
        align="C",
    )
    pdf.ln()

  return bytes(pdf.output())


# Generator Excel Formatted (.xlsx)
def generate_excel_formatted(df):
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Rekap Karyawan"

  header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
  header_fill = PatternFill(
      start_color="1F4E79", end_color="1F4E79", fill_type="solid"
  )
  data_font = Font(name="Calibri", size=10)
  border_thin = Side(border_style="thin", color="D9D9D9")
  border_box = Border(
      left=border_thin, right=border_thin, top=border_thin, bottom=border_thin
  )

  ws.merge_cells("A1:J1")
  ws["A1"] = "LAPORAN DATABASE KARYAWAN"
  ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
  ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

  ws.merge_cells("A2:J2")
  ws["A2"] = (
      f"Tanggal Ekspor: {date.today().strftime('%d-%m-%Y')} | Total Record:"
      f" {len(df)}"
  )
  ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")
  ws.row_dimensions[1].height = 25
  ws.row_dimensions[2].height = 18

  headers = list(df.columns)
  ws.append([])
  ws.append(headers)
  ws.row_dimensions[4].height = 24

  for col_num, _ in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

  for r_idx, row in df.iterrows():
    row_data = list(row)
    ws.append(row_data)
    row_num = ws.max_row
    ws.row_dimensions[row_num].height = 20

    for c_idx in range(1, len(row_data) + 1):
      cell = ws.cell(row=row_num, column=c_idx)
      cell.font = data_font
      cell.border = border_box
      if headers[c_idx - 1] in [
          "ID",
          "Tanggal Bergabung",
          "Akhir Kontrak",
          "Tanggal Resign",
          "Status",
          "Terakhir Diperbarui",
      ]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

  for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
      if cell.row >= 4 and cell.value:
        max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

  output = io.BytesIO()
  wb.save(output)
  return output.getvalue()


# Inisialisasi Data
if "employees" not in st.session_state:
  st.session_state.employees = load_data()


# Helper function ID otomatis
def generate_next_id():
  df = st.session_state.employees
  max_num = 0
  if not df.empty and "ID" in df.columns:
    for emp_id in df["ID"]:
      if str(emp_id).startswith("EMP-"):
        try:
          num = int(str(emp_id).split("-")[1])
          if num > max_num:
            max_num = num
        except ValueError:
          pass
  return f"EMP-{str(max_num + 1).zfill(3)}"


# --- HEADER ---
st.title("Employee Database Manager")
st.caption("Created by iqbalmantam")

# Metric Total Karyawan
df_master_current = st.session_state.employees
total_karyawan = len(df_master_current)
total_aktif = (
    len(df_master_current[df_master_current["Status"] == "Aktif"])
    if "Status" in df_master_current.columns
    else total_karyawan
)
total_resign = (
    len(df_master_current[df_master_current["Status"] == "Resign"])
    if "Status" in df_master_current.columns
    else 0
)

col_m1, col_m2, col_m3 = st.columns(3)
with col_m1:
  st.metric(label="Karyawan Aktif", value=total_aktif)
with col_m2:
  st.metric(label="Karyawan Resign", value=total_resign)
with col_m3:
  st.metric(label="Total Record Data", value=total_karyawan)

st.divider()

# --- SIDEBAR: PILIHAN PERAN (ROLE) ---
st.sidebar.header("🔐 Akses Pengguna")
role = st.sidebar.radio("Pilih Mode Akses:", ["Umum (View Only)", "Administrator"])

is_admin = False

if role == "Administrator":
  pin_input = st.sidebar.text_input("Masukkan PIN Admin:", type="password")
  if pin_input == ADMIN_PIN:
    st.sidebar.success("Akses Administrator Aktif!")
    is_admin = True
  elif pin_input != "":
    st.sidebar.error("PIN Salah!")
  else:
    st.sidebar.info("Masukkan PIN Administrator.")

# --- SIDEBAR: MENU ADMIN ---
if is_admin:
  st.sidebar.markdown("---")
  st.sidebar.header("⚡ Kontrol Administrator")

  if st.sidebar.button("🔄 Sync / Refresh Data"):
    st.session_state.employees = load_data()
    st.rerun()

  # 1. Tambah Karyawan Baru
  with st.sidebar.expander("➕ Tambah Karyawan Baru", expanded=False):
    with st.form("add_employee_form", clear_on_submit=True):
      auto_id = generate_next_id()
      new_id = st.text_input("ID Karyawan", value=auto_id)
      new_name = st.text_input("Nama Lengkap")
      new_role = st.text_input("Posisi")
      new_cc = st.text_input("Cost Center", placeholder="CC-101")
      new_join = st.date_input("Tanggal Bergabung", value=date.today())
      new_end = st.date_input("Akhir Kontrak", value=date.today())
      new_site = st.text_input(
          "Site / Lokasi Kerja", placeholder="Contoh: JDC / Head Office"
      )
      new_status = st.selectbox("Status Karyawan", ["Aktif", "Resign", "PKWT"])

      new_resign_date = "-"
      if new_status == "Resign":
        new_resign_date = st.date_input(
            "Tanggal Resign", value=date.today()
        ).strftime("%Y-%m-%d")

      submit_btn = st.form_submit_button("Simpan Karyawan")
      if submit_btn:
        clean_id = new_id.strip()
        existing_ids = (
            [
                str(x).strip().lower()
                for x in st.session_state.employees["ID"].values
            ]
            if "ID" in st.session_state.employees.columns
            else []
        )

        if not clean_id or not new_name or not new_role or not new_cc:
          st.error("Mohon isi semua kolom yang wajib!")
        elif clean_id.lower() in existing_ids:
          st.error(f"❌ ID '{clean_id}' sudah digunakan!")
        else:
          new_row = {
              "ID": clean_id,
              "Nama Lengkap": new_name,
              "Posisi": new_role,
              "Cost Center": new_cc,
              "Tanggal Bergabung": new_join.strftime("%Y-%m-%d"),
              "Akhir Kontrak": new_end.strftime("%Y-%m-%d"),
              "Tanggal Resign": new_resign_date,
              "Site": new_site,
              "Status": new_status,
              "Terakhir Diperbarui": str(date.today()),
          }
          updated_df = pd.concat(
              [st.session_state.employees, pd.DataFrame([new_row])],
              ignore_index=True,
          )
          save_data(updated_df)
          st.success(f"✅ ID '{clean_id}' berhasil ditambahkan!")
          st.rerun()

  # 2. Bulk Import Data
  with st.sidebar.expander("📥 Import Banyak Data", expanded=False):
    import_type = st.radio(
        "Metode Import:", ["File CSV", "Tempel Teks (Excel/TSV)"]
    )

    if import_type == "File CSV":
      uploaded_file = st.file_uploader("Pilih file CSV", type=["csv"])
      if uploaded_file is not None:
        if st.button("Mulai Import File"):
          try:
            df_import = pd.read_csv(uploaded_file, dtype={"ID": str})
            df_import.columns = [c.strip() for c in df_import.columns]
            if "Jabatan" in df_import.columns:
              df_import.rename(columns={"Jabatan": "Posisi"}, inplace=True)
            if "Status" not in df_import.columns:
              df_import["Status"] = "Aktif"
            if "Tanggal Resign" not in df_import.columns:
              df_import["Tanggal Resign"] = "-"
            df_import["Terakhir Diperbarui"] = str(date.today())

            existing_ids = set(
                str(x).strip().lower()
                for x in st.session_state.employees["ID"].values
            )
            initial_count = len(df_import)

            df_import_filtered = df_import[
                ~df_import["ID"]
                .astype(str)
                .str.strip()
                .str.lower()
                .isin(existing_ids)
            ]
            added_count = len(df_import_filtered)

            if added_count > 0:
              updated_df = pd.concat(
                  [st.session_state.employees, df_import_filtered],
                  ignore_index=True,
              )
              save_data(updated_df)
              st.success(f"Berhasil mengimpor {added_count} data!")
              st.rerun()
            else:
              st.error("Semua ID pada file sudah terdaftar!")
          except Exception as e:
            st.error(f"Gagal membaca file: {e}")

    else:
      pasted_text = st.text_area("Tempel dari Excel", height=150)
      if st.button("Mulai Import Teks"):
        if pasted_text.strip():
          lines = pasted_text.strip().split("\n")
          added_rows = []
          existing_ids = set(
              str(x).strip().lower()
              for x in st.session_state.employees["ID"].values
          )

          for line in lines:
            delimiter = (
                "\t" if "\t" in line else (";" if ";" in line else ",")
            )
            cols = [c.strip() for c in line.split(delimiter)]
            if len(cols) >= 4:
              emp_id, name, role_title, cc = (
                  cols[0],
                  cols[1],
                  cols[2],
                  cols[3],
              )
              join_d = cols[4] if len(cols) > 4 else ""
              end_d = cols[5] if len(cols) > 5 else ""
              resign_d = cols[6] if len(cols) > 6 else "-"
              site_val = cols[7] if len(cols) > 7 else ""
              status_val = cols[8] if len(cols) > 8 else "Aktif"

              if emp_id.lower() not in existing_ids:
                added_rows.append({
                    "ID": emp_id,
                    "Nama Lengkap": name,
                    "Posisi": role_title,
                    "Cost Center": cc,
                    "Tanggal Bergabung": join_d,
                    "Akhir Kontrak": end_d,
                    "Tanggal Resign": resign_d,
                    "Site": site_val,
                    "Status": status_val,
                    "Terakhir Diperbarui": str(date.today()),
                })
                existing_ids.add(emp_id.lower())

          if added_rows:
            updated_df = pd.concat(
                [st.session_state.employees, pd.DataFrame(added_rows)],
                ignore_index=True,
            )
            save_data(updated_df)
            st.success(f"Berhasil menambahkan {len(added_rows)} data baru!")
            st.rerun()

  # 3. KUNCI & HAPUS DATA SNAPSHOT BULANAN
  with st.sidebar.expander("📸 Freeze / Snapshot Bulanan", expanded=False):
    st.subheader("🔒 Simpan Snapshot Baru")
    selected_periode = st.date_input(
        "Pilih Bulan Periode", value=date.today()
    ).strftime("%Y-%m")

    if st.button(f"🔒 Kunci Data {selected_periode}"):
      try:
        df_curr = st.session_state.employees.copy()
        df_active = (
            df_curr[df_curr["Status"] == "Aktif"].copy()
            if "Status" in df_curr.columns
            else df_curr.copy()
        )

        df_active["Periode"] = selected_periode
        df_active["Tanggal Snapshot"] = str(date.today())

        cols_order = [
            "Periode",
            "ID",
            "Nama Lengkap",
            "Posisi",
            "Cost Center",
            "Tanggal Bergabung",
            "Akhir Kontrak",
            "Tanggal Resign",
            "Site",
            "Status",
            "Terakhir Diperbarui",
            "Tanggal Snapshot",
        ]

        df_old_snap = load_snapshot_data()
        if not df_old_snap.empty:
          if "Periode" in df_old_snap.columns:
            df_old_snap = df_old_snap[
                df_old_snap["Periode"] != selected_periode
            ]
          df_new_snap = pd.concat([df_old_snap, df_active[cols_order]])
        else:
          df_new_snap = df_active[cols_order]

        conn.update(worksheet="Snapshot_Bulanan", data=df_new_snap)
        st.success(f"✅ Rekap {selected_periode} berhasil disimpan!")
        st.rerun()
      except Exception as e:
        st.error(f"Gagal melakukan snapshot: {e}")

    # --- FITUR HAPUS SNAPSHOT (CARA 1) ---
    st.markdown("---")
    st.subheader("🗑️ Hapus Snapshot Periode")

    df_snap_exist = load_snapshot_data()
    if not df_snap_exist.empty and "Periode" in df_snap_exist.columns:
      list_snap_periods = sorted(
          df_snap_exist["Periode"].unique(), reverse=True
      )
      period_to_delete = st.selectbox(
          "Pilih Periode yang Ingin Dihapus:", list_snap_periods
      )

      if st.button(f"🗑️ Hapus Snapshot {period_to_delete}"):
        try:
          df_snap_filtered = df_snap_exist[
              df_snap_exist["Periode"] != period_to_delete
          ]
          conn.update(worksheet="Snapshot_Bulanan", data=df_snap_filtered)
          st.success(
              f"✅ Snapshot periode {period_to_delete} berhasil dihapus!"
          )
          st.rerun()
        except Exception as e:
          st.error(f"Gagal menghapus snapshot: {e}")
    else:
      st.info("Belum ada data snapshot untuk dihapus.")

  # Export CSV & Excel
  st.sidebar.markdown("---")
  st.sidebar.subheader("📤 Ekspor Database")

  csv_data = st.session_state.employees.to_csv(index=False).encode("utf-8-sig")
  st.sidebar.download_button(
      label="📄 Ekspor CSV",
      data=csv_data,
      file_name="ekspor_database_karyawan.csv",
      mime="text/csv",
      use_container_width=True,
  )

  excel_data = generate_excel_formatted(st.session_state.employees)
  st.sidebar.download_button(
      label="📊 Ekspor Excel Formatted (.xlsx)",
      data=excel_data,
      file_name=f"Rekap_Karyawan_{date.today().strftime('%Y%m%d')}.xlsx",
      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
      use_container_width=True,
  )


# --- HALAMAN UTAMA ---
if is_admin:
  st.info("🔓 **Mode Akses:** Administrator")
else:
  st.info("👁️ **Mode Akses:** Umum / Guest (View Only)")


# --- DASHBOARD ANALYTICS ---
with st.expander(
    "📊 **Dashboard Analytics & Visualisasi Data**", expanded=True
):
  if not st.session_state.employees.empty:
    df_ana = st.session_state.employees.copy()

    tab_overview, tab_trend, tab_cost = st.tabs([
        "📈 Ringkasan & Status",
        "🗓️ Tren Snapshot Bulanan",
        "💳 Sebaran Cost Center & Site",
    ])

    with tab_overview:
      c1, c2 = st.columns(2)
      with c1:
        if "Status" in df_ana.columns:
          fig_status = px.pie(
              df_ana,
              names="Status",
              title="Komposisi Status Karyawan",
              hole=0.4,
              color_discrete_sequence=px.colors.qualitative.Set2,
          )
          fig_status.update_traces(
              textposition="inside", textinfo="percent+label"
          )
          st.plotly_chart(fig_status, use_container_width=True)

      with c2:
        if "Posisi" in df_ana.columns:
          top_roles = df_ana["Posisi"].value_counts().head(10).reset_index()
          top_roles.columns = ["Posisi", "Jumlah"]
          fig_role = px.bar(
              top_roles,
              x="Jumlah",
              y="Posisi",
              orientation="h",
              title="Top 10 Posisi Terbanyak",
              color="Jumlah",
              color_continuous_scale="Blues",
          )
          fig_role.update_layout(yaxis={"categoryorder": "total ascending"})
          st.plotly_chart(fig_role, use_container_width=True)

    with tab_trend:
      df_snap_hist = load_snapshot_data()
      if not df_snap_hist.empty and "Periode" in df_snap_hist.columns:
        trend_summary = (
            df_snap_hist.groupby("Periode")["ID"]
            .count()
            .reset_index(name="Karyawan Aktif")
        )
        trend_summary = trend_summary.sort_values("Periode")

        fig_trend = px.line(
            trend_summary,
            x="Periode",
            y="Karyawan Aktif",
            markers=True,
            title="Pertumbuhan Jumlah Karyawan Aktif per Periode Snapshot",
            line_shape="spline",
        )
        fig_trend.update_traces(
            line_color="#1F4E79", line_width=3, marker_size=8
        )
        st.plotly_chart(fig_trend, use_container_width=True)
      else:
        st.info("Belum ada data snapshot historis.")

    with tab_cost:
      c3, c4 = st.columns(2)
      with c3:
        if "Cost Center" in df_ana.columns:
          cc_counts = df_ana["Cost Center"].value_counts().reset_index()
          cc_counts.columns = ["Cost Center", "Jumlah"]
          fig_cc = px.bar(
              cc_counts,
              x="Cost Center",
              y="Jumlah",
              title="Jumlah Karyawan per Cost Center",
              color="Jumlah",
              color_continuous_scale="Viridis",
          )
          st.plotly_chart(fig_cc, use_container_width=True)

      with c4:
        if "Site" in df_ana.columns:
          site_counts = (
              df_ana["Site"]
              .replace("", "Belum Diisi")
              .value_counts()
              .reset_index()
          )
          site_counts.columns = ["Site", "Jumlah"]
          fig_site = px.pie(
              site_counts,
              names="Site",
              values="Jumlah",
              title="Distribusi Lokasi Kerja (Site)",
          )
          st.plotly_chart(fig_site, use_container_width=True)
  else:
    st.write("Belum ada data untuk ditampilkan analitiknya.")

st.divider()

# --- FITUR PENCARIAN & FILTER ---
col_mode, col_cat, col_src = st.columns([1.5, 1.5, 3])

with col_mode:
  view_mode = st.selectbox(
      "Tampilkan Data:", ["Master Real-time", "Rekap Snapshot Bulanan"]
  )

df_display = pd.DataFrame()

if view_mode == "Rekap Snapshot Bulanan":
  df_snap_all = load_snapshot_data()
  if not df_snap_all.empty and "Periode" in df_snap_all.columns:
    list_periode = sorted(df_snap_all["Periode"].unique(), reverse=True)
    selected_view_period = st.selectbox("Pilih Periode Rekap:", list_periode)
    df_display = df_snap_all[
        df_snap_all["Periode"] == selected_view_period
    ].copy()
  else:
    st.warning("Belum ada data snapshot yang disimpan.")
    df_display = pd.DataFrame()
else:
  df_display = st.session_state.employees.copy()

with col_cat:
  search_category = st.selectbox(
      "Cari Berdasarkan:",
      [
          "Semua Kolom",
          "Nama Lengkap",
          "Posisi",
          "Cost Center",
          "Site",
          "Status",
      ],
  )

with col_src:
  search_query = st.text_input("🔍 Masukkan kata kunci pencarian...", "")

if search_query and not df_display.empty:
  query = search_query.strip().lower()

  if search_category == "Nama Lengkap":
    df_display = df_display[
        df_display["Nama Lengkap"]
        .astype(str)
        .str.lower()
        .str.contains(query, na=False)
    ]
  elif search_category == "Posisi":
    df_display = df_display[
        df_display["Posisi"].astype(str).str.lower().str.contains(query, na=False)
    ]
  elif search_category == "Cost Center":
    df_display = df_display[
        df_display["Cost Center"]
        .astype(str)
        .str.lower()
        .str.contains(query, na=False)
    ]
  elif search_category == "Site":
    df_display = df_display[
        df_display["Site"].astype(str).str.lower().str.contains(query, na=False)
    ]
  elif search_category == "Status":
    df_display = df_display[
        df_display["Status"].astype(str).str.lower().str.contains(query, na=False)
    ]
  else:
    mask_name = (
        df_display["Nama Lengkap"]
        .astype(str)
        .str.lower()
        .str.contains(query, na=False)
    )
    mask_role = (
        df_display["Posisi"]
        .astype(str)
        .str.lower()
        .str.contains(query, na=False)
        if "Posisi" in df_display.columns
        else False
    )
    mask_cc = (
        df_display["Cost Center"]
        .astype(str)
        .str.lower()
        .str.contains(query, na=False)
    )
    mask_site = (
        df_display["Site"].astype(str).str.lower().str.contains(query, na=False)
    )
    mask_stat = (
        df_display["Status"].astype(str).str.lower().str.contains(query, na=False)
        if "Status" in df_display.columns
        else False
    )
    df_display = df_display[
        mask_name | mask_role | mask_cc | mask_site | mask_stat
    ]

# Header Tabel & Tombol Ekspor/Cetak
col_tb_title, col_pdf_btn = st.columns([3, 1])
with col_tb_title:
  st.subheader(f"📋 Tabel Data Karyawan ({view_mode})")
with col_pdf_btn:
  if not df_display.empty:
    pdf_bytes = generate_pdf(df_display)
    st.download_button(
        label="📄 Cetak / Download PDF",
        data=pdf_bytes,
        file_name=f"Laporan_Karyawan_{date.today().strftime('%Y%m%d')}.pdf",
        mime="application/pdf",
        use_container_width=True,
    )

# Tabel Data
if df_display.empty:
  st.warning("Tidak ada data karyawan yang cocok dengan pencarian.")
else:
  st.dataframe(df_display, use_container_width=True)

# --- EDIT / HAPUS SATUAN ---
if (
    is_admin
    and view_mode == "Master Real-time"
    and not st.session_state.employees.empty
):
  st.divider()
  st.subheader("🛠️ Kelola / Edit / Ubah Status Data Karyawan")

  selected_id = st.selectbox(
      "Pilih ID Karyawan untuk Diubah / Dihapus:",
      options=["-- Pilih ID --"] + list(st.session_state.employees["ID"]),
  )

  if selected_id != "-- Pilih ID --":
    emp_idx = st.session_state.employees[
        st.session_state.employees["ID"] == selected_id
    ].index[0]
    row = st.session_state.employees.loc[emp_idx]

    with st.form("edit_form"):
      st.write(f"Editing: **{row['Nama Lengkap']}** (ID: `{row['ID']}`)")
      e_name = st.text_input("Nama Lengkap", value=row["Nama Lengkap"])
      e_role = st.text_input("Posisi", value=row.get("Posisi", ""))
      e_cc = st.text_input("Cost Center", value=row["Cost Center"])
      e_join = st.text_input(
          "Tanggal Bergabung (YYYY-MM-DD)", value=row["Tanggal Bergabung"]
      )
      e_end = st.text_input(
          "Akhir Kontrak (YYYY-MM-DD)", value=row["Akhir Kontrak"]
      )
      e_site = st.text_input("Site / Lokasi Kerja", value=row.get("Site", ""))

      current_status = row.get("Status", "Aktif")
      status_opts = ["Aktif", "Resign", "PKWT"]
      idx_stat = (
          status_opts.index(current_status)
          if current_status in status_opts
          else 0
      )
      e_status = st.selectbox(
          "Status Karyawan", options=status_opts, index=idx_stat
      )

      e_resign = st.text_input(
          "Tanggal Resign (YYYY-MM-DD)", value=row.get("Tanggal Resign", "-")
      )

      col_save, col_del = st.columns(2)
      with col_save:
        btn_save = st.form_submit_button("💾 Simpan Perubahan")
      with col_del:
        btn_del = st.form_submit_button("🗑️ Hapus Karyawan")

      if btn_save:
        st.session_state.employees.loc[emp_idx, [
            "Nama Lengkap",
            "Posisi",
            "Cost Center",
            "Tanggal Bergabung",
            "Akhir Kontrak",
            "Tanggal Resign",
            "Site",
            "Status",
            "Terakhir Diperbarui",
        ]] = [
            e_name,
            e_role,
            e_cc,
            e_join,
            e_end,
            e_resign,
            e_site,
            e_status,
            str(date.today()),
        ]
        save_data(st.session_state.employees)
        st.success("Data berhasil diperbarui!")
        st.rerun()

      if btn_del:
        updated_df = st.session_state.employees.drop(emp_idx).reset_index(
            drop=True
        )
        save_data(updated_df)
        st.success("Data karyawan berhasil dihapus!")
        st.rerun()
