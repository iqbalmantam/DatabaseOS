import io
from datetime import date
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

def generate_excel_formatted(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Karyawan"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    border_thin = Side(border_style="thin", color="D9D9D9")
    border_box = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    ws.merge_cells("A1:J1")
    ws["A1"] = "LAPORAN DATABASE KARYAWAN"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Tanggal Ekspor: {date.today().strftime('%d-%m-%Y')} | Total Record: {len(df)}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18

    headers = list(df.columns)
    ws.append([])
    ws.append(headers)
    ws.row_dimensions[4].height = 24

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for _, row in df.iterrows():
        row_data = list(row)
        ws.append(row_data)
        row_num = ws.max_row
        ws.row_dimensions[row_num].height = 20

        for c_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.font = data_font
            cell.border = border_box
            if headers[c_idx - 1] in ["ID", "Tanggal Bergabung", "Akhir Kontrak", "Tanggal Resign", "Status", "Terakhir Diperbarui"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= 4 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
